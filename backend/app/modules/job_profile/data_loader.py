from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any


SUPPORTED_SUFFIXES = {".csv", ".json", ".xls", ".xlsx"}
REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_RAW_DATA_DIR = REPO_ROOT / "data" / "raw" / "official"


def detect_source_format(file_path: str | Path) -> str:
    suffix = Path(file_path).suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError(f"Unsupported source format: {suffix}")
    return suffix.lstrip(".")



def resolve_input_path(input_path: str | Path | None = None, raw_data_dir: str | Path | None = None) -> Path:
    if input_path:
        path = Path(input_path)
        if path.exists():
            return path.resolve()
        repo_relative = REPO_ROOT / path
        if repo_relative.exists():
            return repo_relative.resolve()
        raise FileNotFoundError(f"Job source file was not found: {input_path}")
    return discover_latest_source_file(raw_data_dir)



def discover_latest_source_file(raw_data_dir: str | Path | None = None) -> Path:
    target_dir = Path(raw_data_dir) if raw_data_dir else DEFAULT_RAW_DATA_DIR
    if not target_dir.is_absolute():
        repo_relative = REPO_ROOT / target_dir
        if repo_relative.exists():
            target_dir = repo_relative
    if not target_dir.exists():
        raise FileNotFoundError(f"Raw job data directory does not exist: {target_dir}")

    candidates = [
        path for path in target_dir.iterdir() if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES
    ]
    if not candidates:
        raise FileNotFoundError(f"No supported job source files were found in: {target_dir}")
    return max(candidates, key=lambda item: item.stat().st_mtime).resolve()



def _safe_scalar(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    return str(value)



def _normalize_record(record: dict[str, Any]) -> dict[str, Any]:
    return {str(key).strip(): _safe_scalar(value) for key, value in record.items()}



def _read_csv_records(file_path: Path) -> list[dict[str, Any]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with file_path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                return [_normalize_record(row) for row in reader]
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    return []



def _read_json_records(file_path: Path) -> list[dict[str, Any]]:
    payload = json.loads(file_path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return [_normalize_record(row) for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        if isinstance(payload.get("data"), list):
            return [_normalize_record(row) for row in payload["data"] if isinstance(row, dict)]
        return [_normalize_record(payload)]
    raise ValueError("JSON payload must be an object or an array of objects")



def _tabular_rows_to_records(rows: list[tuple[Any, ...]] | list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        values = {headers[index]: _safe_scalar(value) for index, value in enumerate(row) if index < len(headers)}
        if any(value not in (None, "") for value in values.values()):
            records.append(_normalize_record(values))
    return records



def _read_xlsx_records(file_path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Reading .xlsx files requires openpyxl") from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        return _tabular_rows_to_records(rows)
    finally:
        workbook.close()



def _read_xls_records(file_path: Path) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Reading .xls files requires xlrd") from exc

    workbook = xlrd.open_workbook(file_path.as_posix())
    sheet = workbook.sheet_by_index(0)
    rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
    return _tabular_rows_to_records(rows)



def load_source_records(file_path: str | Path) -> tuple[str, list[dict[str, Any]]]:
    path = resolve_input_path(file_path)
    source_format = detect_source_format(path)
    if source_format == "csv":
        records = _read_csv_records(path)
    elif source_format == "json":
        records = _read_json_records(path)
    elif source_format == "xls":
        records = _read_xls_records(path)
    else:
        records = _read_xlsx_records(path)
    if not records:
        raise ValueError(f"No job records were found in {path}")
    return source_format, records



def normalize_field_name(field_name: str) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", field_name.strip().lower())

